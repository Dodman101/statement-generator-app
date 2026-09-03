"""
Statement generation routes - FastAPI port of the original Flask blueprint.

Background job model: StatementGenerator.run() is pure blocking code
(subprocess calls to LibreOffice, synchronous file I/O, docxtpl rendering)
and stays exactly that way - rewriting it to be "async" wouldn't make any
of that actually non-blocking, since LibreOffice itself is a subprocess, not
an async-aware library. Instead, run_generation_job() is a coroutine that
hands the blocking work to a dedicated ThreadPoolExecutor via
loop.run_in_executor(), then awaits the (now-async) database update once
that returns. This keeps the "at most MAX_CONCURRENT_JOBS concurrent jobs"
concurrency limit from the original design, while letting the DB calls use
asyncpg properly instead of blocking the event loop.

Memory: nothing here is unbounded per se, but a few things scale with job
size/concurrency and are worth knowing about if RSS climbs under load -
see the comments by MAX_CONCURRENT_JOBS, LIBREOFFICE_CONCURRENCY, and
_release_memory_to_os() below for the specifics and the fixes in place.
"""
import os
import re
import gc
import ctypes
import signal
import shutil
import subprocess
import platform
import logging
import uuid
import zipfile
import asyncio
from datetime import datetime, timedelta
from threading import Event, Thread, Semaphore
from concurrent.futures import ThreadPoolExecutor

import openpyxl
from docxtpl import DocxTemplate
from PyPDF2 import PdfWriter, PdfReader
from werkzeug.utils import secure_filename

from fastapi import APIRouter, Request, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel

from app.auth import require_api_key
from app.db import create_job, update_job_status, count_jobs_this_month, get_connection
from app.observability import capture_job_failure, capture_service_error

logger = logging.getLogger(__name__)

router = APIRouter(tags=["statement-generator"])

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # app/
TEMPLATES_DIR = os.path.join(BASE_DIR, 'templates')
STATIC_DIR = os.path.join(BASE_DIR, 'static')
templates = Jinja2Templates(directory=TEMPLATES_DIR)

ALLOWED_TEMPLATE_EXTENSIONS = {'docx', 'doc'}
ALLOWED_DATA_EXTENSIONS = {'xlsx', 'xls'}

# Temporary downloads and progress tracking.
# NOTE: these live in process memory. That's fine as long as the app runs as
# a single worker process (e.g. `uvicorn main:app --workers 1`). If you ever
# scale to multiple worker processes, a request can land on a worker that
# never ran the job and progress/download lookups will 404 - move this to
# Redis (or similar) before doing that.
TEMP_DOWNLOADS = {}
PROGRESS_STATUS = {}

CONVERSION_BATCH_TARGET_BYTES = int(os.environ.get(
    "CONVERSION_BATCH_TARGET_BYTES", str(8 * 1024 * 1024)))  # ~8MB of generated .docx per LibreOffice call
# Hard ceiling on doc count per batch regardless of size, so a template
# that renders to tiny files doesn't produce a single 2000-file command
# line. Hard floor is implicitly 1 - see _AdaptiveBatcher below.
CONVERSION_MAX_BATCH_SIZE = int(os.environ.get("CONVERSION_MAX_BATCH_SIZE", "30"))
# How many times a failing batch is retried, at its original size, before
# we give up on it as a whole and start bisecting to isolate the bad
# file(s) - see StatementGenerator._convert_with_retries.
BATCH_MAX_RETRIES = int(os.environ.get("CONVERSION_BATCH_MAX_RETRIES", "2"))
# After this many *top-level* batches in a row fail completely (all
# retries exhausted at their original size, before any bisection), stop
# retrying/bisecting further and abort the job instead. Several unrelated
# batches failing outright in a row is a much stronger signal of a
# systemic problem - LibreOffice unhealthy, disk full, the environment
# itself broken - than of several unrelated bad documents, and grinding
# through retries-then-bisection for every remaining batch would just
# multiply timeouts without ever succeeding.
CIRCUIT_BREAKER_THRESHOLD = int(os.environ.get("CONVERSION_CIRCUIT_BREAKER_THRESHOLD", "2"))

# How many full jobs (docx generation + PDF conversion + password
# protection) can run at once. Configurable via env var so it can be tuned
# down on a small Render instance without a redeploy.
MAX_CONCURRENT_JOBS = int(os.environ.get("MAX_CONCURRENT_JOBS", "3"))

# LibreOffice is the real memory hog here, not our own Python code - each
# `soffice --headless` invocation can hold 150-300MB+ RSS on its own. The
# ThreadPoolExecutor below caps how many *jobs* run at once, but every one
# of those jobs still calls into LibreOffice independently, so at
# MAX_CONCURRENT_JOBS=3 you could still get 3 soffice processes launching in
# the same instant. This semaphore caps concurrent LibreOffice subprocesses
# specifically, separate from (and normally tighter than) job concurrency,
# so conversions from different jobs queue up instead of all running at once.
LIBREOFFICE_CONCURRENCY = int(os.environ.get("LIBREOFFICE_CONCURRENCY", "2"))
libreoffice_semaphore = Semaphore(LIBREOFFICE_CONCURRENCY)

# Dedicated thread pool for the blocking generation work (LibreOffice +
# docxtpl). Separate from FastAPI/uvicorn's own thread pool so job
# concurrency is a deliberate, bounded number, not whatever uvicorn happens
# to be running.
executor = ThreadPoolExecutor(max_workers=MAX_CONCURRENT_JOBS)


def _release_memory_to_os():
    """Force Python to reclaim garbage and, on Linux, actually hand freed
    heap memory back to the OS.

    CPython's allocator (and glibc's malloc arenas underneath it) will
    happily hold onto memory that's been freed rather than returning it -
    this is normal behavior, not a leak, but across a long-running worker
    process that repeatedly creates and discards big short-lived objects
    (one DocxTemplate/openpyxl/PdfReader per statement, over and over) it's
    exactly what makes RSS climb during a big job and never come back down
    afterwards. gc.collect() clears any reference cycles (lxml, which
    docxtpl/python-docx sit on top of, is a common source of these) so
    CPython's refcounting alone won't free them promptly; malloc_trim(0)
    then asks glibc to actually release the freed arenas back to the OS
    instead of keeping them reserved for next time. Call this after each
    memory-heavy phase of a job, not on every row - it has real cost.
    """
    gc.collect()
    if platform.system() == 'Linux':
        try:
            ctypes.CDLL("libc.so.6").malloc_trim(0)
        except Exception:
            pass


def _current_rss_mb():
    """Current resident memory of this whole process, in MB.

    Reads /proc/self/status rather than resource.getrusage().ru_maxrss,
    which is a high-water mark that only ever goes up - it can't tell you
    whether a given phase actually released memory afterward. Returns None
    off Linux or if /proc isn't readable (e.g. sandboxed environments).
    """
    try:
        with open('/proc/self/status') as f:
            for line in f:
                if line.startswith('VmRSS:'):
                    return int(line.split()[1]) / 1024  # kB -> MB
    except Exception:
        return None
    return None


def _log_memory(temp_id, stage):
    """Log this process's current memory next to a job/stage label, so a
    Render log search for a temp_id shows exactly which phase memory grew
    or shrank at - actual numbers instead of the earlier guesswork."""
    rss = _current_rss_mb()
    if rss is not None:
        logger.info(f"[{temp_id}] memory after {stage}: {rss:.1f} MB RSS")


def _reap_orphaned_libreoffice_processes():
    """Safety net: kill any soffice/soffice.bin process still running for a
    job this process no longer has any record of (completed, errored, or
    never fully tracked).

    Under normal operation, _kill_libreoffice_process_group() below should
    prevent orphans from ever existing - but this catches anything that
    slips through it (a hard crash, an OOM-killer that took the wrapper but
    missed a forked child, a deploy that restarted mid-job) before it can
    pile up and eat memory or contend with new conversions for resources.
    Runs on the same interval as cleanup_expired_downloads().
    """
    if platform.system() != 'Linux':
        return
    try:
        for pid_entry in os.listdir('/proc'):
            if not pid_entry.isdigit():
                continue
            try:
                with open(f'/proc/{pid_entry}/cmdline', 'rb') as f:
                    cmdline = f.read().decode(errors='ignore')
            except (FileNotFoundError, ProcessLookupError, PermissionError):
                continue
            if 'soffice' not in cmdline or 'lo_profiles' not in cmdline:
                continue
            match = re.search(r'lo_profiles/([0-9a-fA-F-]{36})', cmdline)
            if not match:
                continue
            temp_id = match.group(1)
            if temp_id in TEMP_DOWNLOADS:
                continue  # job still tracked/active - leave it running
            try:
                pid = int(pid_entry)
                pgid = os.getpgid(pid)
                os.killpg(pgid, signal.SIGKILL)
                logger.warning(f"Reaped orphaned LibreOffice process (pid={pid}) for untracked job {temp_id}")
            except (ProcessLookupError, PermissionError):
                pass
    except Exception as e:
        logger.error(f"Error while reaping orphaned LibreOffice processes: {e}")


# ---------------------------------------------------------------------------
# Response models - so /docs shows real schemas, not untyped dicts.
# ---------------------------------------------------------------------------

class ProcessStatementResponse(BaseModel):
    temp_id: str
    message: str
    status_url: str
    download_link: str


class ProgressResponse(BaseModel):
    state: str
    message: str
    progress: int
    warnings: list[str] | None = None


class HealthResponse(BaseModel):
    status: str
    message: str | None = None


def _extension(filename: str) -> str:
    return filename.rsplit('.', 1)[1].lower() if '.' in filename else ''


def allowed_template(filename: str) -> bool:
    return _extension(filename) in ALLOWED_TEMPLATE_EXTENSIONS


def allowed_data_file(filename: str) -> bool:
    return _extension(filename) in ALLOWED_DATA_EXTENSIONS


def _client_root(client_id) -> str:
    """The filesystem root that belongs to one client, and nobody else.
    Every job folder for this client lives under here - never directly
    under the shared statement_generator/ directory."""
    base_folder = '/tmp' if platform.system() == 'Linux' else BASE_DIR
    return os.path.join(base_folder, 'statement_generator', f'client_{client_id}')


def _safe_join(root: str, *parts) -> str:
    """Join path components under `root` and refuse to return anything outside
    it. This isn't a check bolted onto one call site - it's the only function
    in this codebase allowed to build a job's on-disk path, so "does this path
    belong to this client" is enforced by construction rather than by every
    caller remembering to check."""
    root = os.path.realpath(root)
    candidate = os.path.realpath(os.path.join(root, *parts))
    if candidate != root and not candidate.startswith(root + os.sep):
        raise ValueError(f"Refusing to build a path outside its owner's root: {candidate}")
    return candidate


def create_temp_folder(client_id, temp_id: str) -> str:
    """Create this job's folder, nested under this client's own root."""
    folder_path = _safe_join(_client_root(client_id), temp_id)
    os.makedirs(folder_path, exist_ok=True)
    return folder_path


def cleanup_expired_downloads():
    """Remove expired temporary downloads."""
    current_time = datetime.utcnow()
    expired_ids = [temp_id for temp_id, data in list(TEMP_DOWNLOADS.items()) if data['expiry'] < current_time]

    for temp_id in expired_ids:
        download_info = TEMP_DOWNLOADS.pop(temp_id, None)
        PROGRESS_STATUS.pop(temp_id, None)
        if download_info:
            try:
                shutil.rmtree(download_info['folder'], ignore_errors=True)
                logger.info(f"Cleaned up expired files for temp_id: {temp_id}")
                client_root = os.path.dirname(download_info['folder'])
                if os.path.isdir(client_root) and not os.listdir(client_root):
                    os.rmdir(client_root)
            except Exception as e:
                logger.error(f"Error cleaning up temp_id {temp_id}: {e}")


cleanup_stop_event = Event()


def schedule_cleanup():
    while not cleanup_stop_event.is_set():
        cleanup_expired_downloads()
        _reap_orphaned_libreoffice_processes()
        # Shorter interval than before (was 600s) - the reaper is the
        # safety net for orphaned LibreOffice processes, and those cost
        # real memory for every minute they're left running.
        cleanup_stop_event.wait(120)


# Plain daemon thread, not asyncio - this loop doesn't touch the DB or need
# to run on the event loop, and starting it here (module import time) keeps
# it identical to the original design.
Thread(target=schedule_cleanup, daemon=True).start()


def set_progress(temp_id, state, message, progress, warnings=None):
    entry = {"state": state, "message": message, "progress": progress}
    if warnings:
        entry["warnings"] = warnings
    PROGRESS_STATUS[temp_id] = entry


class StatementGenerator:
    """Unchanged from the Flask version - this is pure blocking code with no
    framework dependency at all, so there's nothing to port here."""

    ID_HEADER_NAMES = {'id', 'client id', 'member id', 'id number'}
    NAME_HEADER_NAMES = {'name', 'client name', 'member name'}

    def __init__(self, template_path, data_path, output_folder, temp_id):
        self.template_path = template_path
        self.data_path = data_path
        self.output_folder = output_folder
        self.temp_id = temp_id
        # read_only streams rows from the XML instead of materializing every
        # cell in memory up front - for a data file with a few thousand rows
        # that's the difference between a small, constant footprint and one
        # that scales with file size before a single statement is generated.
        self.workbook = openpyxl.load_workbook(data_path, read_only=True, data_only=True)
        self.sheet = self.workbook.active
        self.header_row = next(self.sheet.iter_rows(values_only=True))
        self.rows_by_safe_id = {}
        self.unprotected_names = []
        self.failed_statements = []
        self.consecutive_batch_failures = 0
        self._last_batch_error = None
        self.libreoffice_path = self._get_libreoffice_path()

        self.id_column_index = self._find_column(self.ID_HEADER_NAMES)
        self.name_column_index = self._find_column(self.NAME_HEADER_NAMES)
        if self.id_column_index is None:
            raise ValueError("No 'ID' column found in the Excel headers (expected one of: id, client id, member id, id number).")
        if self.name_column_index is None:
            raise ValueError("No 'name' column found in the Excel headers (expected one of: name, client name, member name).")

    def _find_column(self, candidate_names):
        for index, header in enumerate(self.header_row):
            if header and header.strip().lower() in candidate_names:
                return index
        return None

    def _get_libreoffice_path(self):
        if platform.system() == 'Linux':
            linux_paths = [
                '/usr/bin/soffice',
                '/usr/bin/libreoffice',
                '/usr/lib/libreoffice/program/soffice',
                '/opt/libreoffice*/program/soffice'
            ]
            for path in linux_paths:
                if '*' in path:
                    import glob
                    matching_paths = glob.glob(path)
                    if matching_paths:
                        return matching_paths[0]
                elif os.path.exists(path):
                    return path
            try:
                result = subprocess.run(['which', 'soffice'], capture_output=True, text=True)
                if result.returncode == 0 and result.stdout.strip():
                    return result.stdout.strip()
            except subprocess.SubprocessError:
                pass
            snap_path = '/snap/bin/libreoffice'
            if os.path.exists(snap_path):
                return snap_path
            logger.warning("LibreOffice not found in common locations. Please install it using: sudo apt-get install libreoffice")
            return None
        elif platform.system() == 'Windows':
            for path in (r'C:\Program Files\LibreOffice\program\soffice.exe',
                         r'C:\Program Files (x86)\LibreOffice\program\soffice.exe'):
                if os.path.exists(path):
                    return path
        return 'soffice'

    def format_number(self, value):
        if isinstance(value, (int, float)):
            return "-" if value == 0 else "{:,.0f}".format(value)
        return str(value) if value is not None else "-"

    def validate_template(self):
        template = DocxTemplate(self.template_path)
        placeholders = template.get_undeclared_template_variables()
        available_fields = {h.strip().replace(' ', '_') for h in self.header_row if h}
        missing_fields = [field for field in placeholders if field not in available_fields]
        if missing_fields:
            raise ValueError(f"Template fields not found in Excel headers: {', '.join(sorted(missing_fields))}")

    def generate_and_convert(self):
        """Generate each statement's .docx and convert it to PDF in
        batches, interleaved row-by-row instead of generating every
        document up front and only then converting.

        Previously all N documents were written to disk before conversion
        started at all, so a job's peak disk usage scaled with its total
        row count - for a couple thousand rows with a heavy template
        (embedded images, big tables) that could add up to gigabytes of
        ephemeral disk before a single PDF existed. Flushing a batch to
        LibreOffice as soon as it's ready bounds peak disk usage to
        roughly one batch's worth of documents, regardless of job size.

        Batch size is adaptive by *size*, not row count: rows accumulate
        into a batch until either their generated .docx bytes reach
        CONVERSION_BATCH_TARGET_BYTES or the count hits
        CONVERSION_MAX_BATCH_SIZE, whichever comes first. A simple
        template produces small .docx files and gets bigger batches
        (fewer, faster LibreOffice invocations); a complex template with
        embedded images produces bigger files and gets smaller batches
        automatically, which keeps any single LibreOffice call less
        likely to run long enough to hit its timeout.
        """
        set_progress(self.temp_id, "processing", "Generating and converting statements...", 5)
        total_rows = max(self.sheet.max_row - 1, 0)
        if total_rows == 0:
            raise ValueError("The data file has no data rows below the header.")

        if not self.libreoffice_path:
            raise RuntimeError("LibreOffice not found on the server. Install it with: sudo apt-get install libreoffice")

        env = os.environ.copy()
        if platform.system() == 'Linux':
            env['HOME'] = '/tmp'

        profile_dir = os.path.join('/tmp', 'lo_profiles', self.temp_id)
        os.makedirs(profile_dir, exist_ok=True)
        user_installation_arg = f'-env:UserInstallation=file://{profile_dir}'

        current_batch = []
        current_batch_bytes = 0
        processed = 0

        def flush_batch():
            nonlocal current_batch, current_batch_bytes
            if not current_batch:
                return
            # Handles its own retries and, on persistent failure, isolates
            # and skips the specific bad file(s) instead of raising - see
            # _convert_with_retries for the policy.
            self._convert_with_retries(current_batch, env, user_installation_arg)
            current_batch = []
            current_batch_bytes = 0

        try:
            for index, row in enumerate(self.sheet.iter_rows(min_row=2, values_only=True), start=1):
                if not row or all(v is None for v in row):
                    continue

                client_id = row[self.id_column_index]
                client_name = row[self.name_column_index]
                if client_name is None or str(client_name).strip() == '':
                    raise ValueError(f"Row {index + 1} is missing a value in the name column.")

                id_missing = client_id is None or str(client_id).strip() == ''
                if id_missing:
                    name_part = re.sub(r'[^a-zA-Z0-9_-]', '_', str(client_name).strip())[:30]
                    safe_id = f"row{index}_{name_part}"
                else:
                    safe_id = re.sub(r'[^a-zA-Z0-9_-]', '_', str(client_id).strip())

                if safe_id in self.rows_by_safe_id:
                    safe_id = f"{safe_id}_{index}"

                self.rows_by_safe_id[safe_id] = {
                    "id": None if id_missing else str(client_id).strip(),
                    "name": str(client_name).strip(),
                    "id_missing": id_missing,
                }

                template = DocxTemplate(self.template_path)
                context = {
                    header.strip().replace(' ', '_'): self.format_number(value)
                    for header, value in zip(self.header_row, row) if header
                }

                output_path = os.path.join(self.output_folder, f"output_{safe_id}.docx")
                try:
                    template.render(context)
                    template.save(output_path)
                except Exception as e:
                    raise RuntimeError(f"Failed to generate document for row {index + 1} ({client_name}): {e}")

                current_batch.append(output_path)
                current_batch_bytes += os.path.getsize(output_path)
                processed += 1

                if (current_batch_bytes >= CONVERSION_BATCH_TARGET_BYTES
                        or len(current_batch) >= CONVERSION_MAX_BATCH_SIZE):
                    flush_batch()

                # 5-80% of the progress bar covers this combined phase now
                # that generation and conversion happen together per batch.
                progress_percent = int((processed / total_rows) * 75) + 5
                set_progress(self.temp_id, "processing", f"Processing statement {processed}/{total_rows}...", progress_percent)

                # DocxTemplate/lxml objects can form reference cycles that
                # refcounting alone won't clean up promptly, so on a job
                # with thousands of rows, memory can climb steadily through
                # this loop even though each `template` is discarded every
                # iteration. Nudge the collector periodically, not per row.
                if processed % 200 == 0:
                    gc.collect()

            flush_batch()  # whatever's left under the size/count threshold
        finally:
            shutil.rmtree(profile_dir, ignore_errors=True)
            # The workbook has done its job (every row has been read and
            # turned into a PDF) - drop it now instead of holding it for
            # the rest of the run, which still has password protection
            # ahead of it.
            self.workbook.close()
            self.workbook = None
            self.sheet = None
            _release_memory_to_os()
            _log_memory(self.temp_id, "generate+convert")

        set_progress(self.temp_id, "processing", "Documents generated and converted.", 80)

    def _try_convert_batch(self, batch, env, user_installation_arg):
        """Attempt to convert `batch` at its current size, retrying
        transient failures up to BATCH_MAX_RETRIES times. Returns True and
        removes the source .docx files on success. Returns False (leaving
        the source files in place for the caller to bisect or clean up) if
        every attempt at this size failed; the failure is stashed on
        self._last_batch_error for the caller to report.
        """
        cmd = [self.libreoffice_path, '--headless', user_installation_arg, '--convert-to', 'pdf',
               '--outdir', self.output_folder] + batch

        for attempt in range(1, BATCH_MAX_RETRIES + 2):  # +1 for the initial try
            try:
                # Only LIBREOFFICE_CONCURRENCY of these run at once across
                # ALL jobs in this process, regardless of how many jobs the
                # thread pool is otherwise running - see the comment by the
                # semaphore's definition.
                with libreoffice_semaphore:
                    self._run_libreoffice(cmd, env)

                missing = [f for f in batch if not os.path.exists(os.path.splitext(f)[0] + '.pdf')]
                if missing:
                    raise RuntimeError(
                        f"PDF was not created for {', '.join(os.path.basename(m) for m in missing)}.")

                for letter_file in batch:
                    os.remove(letter_file)
                return True
            except RuntimeError as e:
                self._last_batch_error = e
                logger.warning(
                    f"[{self.temp_id}] batch of {len(batch)} document(s) failed "
                    f"(attempt {attempt}/{BATCH_MAX_RETRIES + 1}): {e}")

        return False

    def _convert_with_retries(self, batch, env, user_installation_arg, top_level=True):
        """Convert `batch` (a list of .docx paths) to PDF, retrying
        transient failures and isolating a persistently bad file instead
        of failing the whole job over it.

        Policy: retry the batch at its current size up to BATCH_MAX_RETRIES
        times (handles a flaky/slow LibreOffice invocation - the common
        case). If it still won't convert and has more than one file, split
        it in half and recurse on each half - this narrows down which
        file(s) are actually the problem rather than discarding an entire
        batch because of one bad document. A batch of exactly one file that
        still fails after retries is recorded as a failed statement and
        skipped; everything else in the job keeps going.

        `top_level` marks calls made directly from the main batching loop,
        as opposed to bisected sub-batches - only those count toward the
        circuit breaker below, since that's what tells us whether *whole
        batches* keep failing outright (systemic) rather than isolated
        single files within an otherwise-healthy batch (expected, and what
        bisection exists to handle).
        """
        if self._try_convert_batch(batch, env, user_installation_arg):
            if top_level:
                self.consecutive_batch_failures = 0
            return

        if top_level:
            self.consecutive_batch_failures += 1
            if self.consecutive_batch_failures >= CIRCUIT_BREAKER_THRESHOLD:
                raise RuntimeError(
                    f"{self.consecutive_batch_failures} batches in a row failed to convert even "
                    f"after retries - this looks like a systemic problem (LibreOffice, disk, or "
                    f"the environment itself) rather than a bad file, so the job is stopping "
                    f"instead of working through every remaining statement one at a time. "
                    f"Last error: {self._last_batch_error}"
                )

        if len(batch) > 1:
            mid = len(batch) // 2
            self._convert_with_retries(batch[:mid], env, user_installation_arg, top_level=False)
            self._convert_with_retries(batch[mid:], env, user_installation_arg, top_level=False)
            return

        # A single file that still won't convert - skip it and move on
        # rather than losing everything else the job already produced.
        failed_path = batch[0]
        safe_id = self._safe_id_from_filename(os.path.basename(failed_path))
        row_info = self.rows_by_safe_id.get(safe_id) if safe_id else None
        name = row_info['name'] if row_info else os.path.basename(failed_path)
        self.failed_statements.append(name)
        logger.error(
            f"[{self.temp_id}] giving up on '{name}' after {BATCH_MAX_RETRIES + 1} attempts: {self._last_batch_error}")
        try:
            os.remove(failed_path)
        except OSError:
            pass

    def _run_libreoffice(self, cmd, env):
        """Run one LibreOffice conversion batch, making sure that if it
        times out (or fails) we kill LibreOffice's *entire* process tree -
        not just the one PID we launched.

        `soffice` is a wrapper that forks the real worker, `soffice.bin`.
        subprocess.run(..., timeout=...) only ever signals the PID it
        started, so on timeout the wrapper dies but soffice.bin - the
        process actually holding the memory - is orphaned and keeps
        running. That's a likely cause of both symptoms seen in
        production: orphaned soffice.bin processes accumulate and eat
        memory over time, and once several are alive, new conversions have
        to contend with them for memory/CPU, which shows up as exactly the
        kind of timeout that was hit here - even on a small first batch.

        Starting the process in its own session (start_new_session=True)
        puts the wrapper and everything it forks into one process group,
        so on timeout or failure we can reliably kill the whole group with
        a single signal instead of leaving children behind.
        """
        proc = subprocess.Popen(
            cmd, env=env, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
            start_new_session=True,
        )
        try:
            _, stderr = proc.communicate(timeout=180)
        except subprocess.TimeoutExpired:
            self._kill_libreoffice_process_group(proc)
            raise RuntimeError("PDF conversion timed out. Try again with a smaller batch of statements.")

        if proc.returncode != 0:
            self._kill_libreoffice_process_group(proc)  # in case it left children behind
            raise RuntimeError(f"PDF conversion failed: {stderr.decode(errors='ignore')}")

    @staticmethod
    def _kill_libreoffice_process_group(proc):
        try:
            pgid = os.getpgid(proc.pid)
            os.killpg(pgid, signal.SIGKILL)
        except ProcessLookupError:
            pass  # already gone
        except Exception as e:
            logger.warning(f"Could not clean up LibreOffice process group for pid {proc.pid}: {e}")
        try:
            proc.wait(timeout=5)
        except Exception:
            pass

    def _safe_id_from_filename(self, filename):
        """Recover the safe_id from an `output_<safe_id>.<ext>` filename,
        regardless of extension - used for both the .docx files handled
        during conversion and the .pdf files handled afterward."""
        if filename.startswith("output_"):
            return os.path.splitext(filename)[0][len("output_"):]
        return None

    def apply_password_protection(self):
        set_progress(self.temp_id, "processing", "Applying password protection...", 88)
        for count, filename in enumerate(os.listdir(self.output_folder), start=1):
            if not filename.endswith(".pdf"):
                continue
            safe_id = self._safe_id_from_filename(filename)
            row_info = self.rows_by_safe_id.get(safe_id) if safe_id else None
            if not row_info:
                logger.warning(f"No client record found for '{filename}'; skipping password protection for this file.")
                continue
            if row_info.get('id_missing'):
                logger.warning(f"Skipping password protection for '{row_info['name']}' - no ID was provided for this row.")
                self.unprotected_names.append(row_info['name'])
                continue

            full_path = os.path.join(self.output_folder, filename)
            reader = PdfReader(full_path)
            writer = PdfWriter()
            for page in reader.pages:
                writer.add_page(page)
            writer.encrypt(row_info['id'])
            with open(full_path, 'wb') as protected_file:
                writer.write(protected_file)

            if count % 200 == 0:
                gc.collect()

        _log_memory(self.temp_id, "password protection")

    def rename_pdfs(self):
        set_progress(self.temp_id, "processing", "Renaming files...", 95)
        for filename in os.listdir(self.output_folder):
            if not filename.endswith(".pdf"):
                continue
            safe_id = self._safe_id_from_filename(filename)
            row_info = self.rows_by_safe_id.get(safe_id) if safe_id else None
            old_path = os.path.join(self.output_folder, filename)
            if not row_info:
                logger.warning(f"Could not match '{filename}' back to a client record; leaving filename as-is.")
                continue

            sanitized_name = re.sub(r'[<>:"/\\|?*]', '_', row_info['name'])
            new_name = f"{sanitized_name}.pdf"
            new_path = os.path.join(self.output_folder, new_name)
            if os.path.exists(new_path) and new_path != old_path:
                base, ext = os.path.splitext(new_name)
                new_name = f"{base}_{safe_id}{ext}"
                new_path = os.path.join(self.output_folder, new_name)
            shutil.move(old_path, new_path)

    def run(self, password_protection=False):
        _log_memory(self.temp_id, "job start")
        self.validate_template()
        self.generate_and_convert()
        if password_protection:
            self.apply_password_protection()
        self.rename_pdfs()

        warnings = []
        if password_protection and self.unprotected_names:
            warnings.extend(
                f"'{name}' was generated without password protection because no ID was provided for that row."
                for name in self.unprotected_names
            )
        if self.failed_statements:
            warnings.extend(
                f"'{name}' could not be converted to PDF after repeated attempts and was skipped."
                for name in self.failed_statements
            )
        warnings = warnings or None

        done_message = "Done - your statements are ready."
        if self.failed_statements:
            done_message = (
                f"Done, with {len(self.failed_statements)} statement(s) skipped - see warnings."
            )
        set_progress(self.temp_id, "completed", done_message, 100, warnings=warnings)
        # Last chance to hand this job's memory back to the OS before the
        # worker thread picks up (or waits for) the next one.
        _release_memory_to_os()
        _log_memory(self.temp_id, "job end")


async def run_generation_job(temp_id, client_id, client_label, template_path, data_path, output_folder, password_protection):
    """The blocking generation work runs in the dedicated executor; the
    surrounding coroutine handles the async DB update and cleanup."""
    loop = asyncio.get_running_loop()

    def _run_sync():
        generator = StatementGenerator(template_path, data_path, output_folder, temp_id)
        generator.run(password_protection=password_protection)

    try:
        await loop.run_in_executor(executor, _run_sync)
        await update_job_status(temp_id, 'completed', client_id)
    except Exception as e:
        set_progress(temp_id, "error", str(e), 0)
        _release_memory_to_os()
        _log_memory(temp_id, f"job error ({e.__class__.__name__})")
        capture_job_failure(e, temp_id=temp_id, client_label=client_label, client_id=client_id)
        try:
            await update_job_status(temp_id, 'error', client_id)
        except Exception as db_err:
            logger.error(f"Could not record job error in database for {temp_id}: {db_err}")
    finally:
        for path in (template_path, data_path):
            try:
                if os.path.exists(path):
                    os.remove(path)
            except Exception as e:
                logger.warning(f"Could not remove temporary upload {path}: {e}")


def count_data_rows(data_path):
    workbook = openpyxl.load_workbook(data_path, read_only=True)
    try:
        sheet = workbook.active
        count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and any(value is not None for value in row):
                count += 1
        return count
    finally:
        workbook.close()


def generate_temp_id():
    temp_id = str(uuid.uuid4())
    expiry_time = datetime.utcnow() + timedelta(hours=2)
    return temp_id, expiry_time


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@router.get("/health", response_model=HealthResponse, summary="Service health check")
async def health():
    """Public, unauthenticated - for uptime monitoring. Checks the one
    dependency that actually matters: can we reach the database."""
    try:
        async with get_connection() as conn:
            await conn.execute("SELECT 1")
        return HealthResponse(status="ok")
    except Exception as e:
        capture_service_error(e, where="health_check")
        raise HTTPException(status_code=503, detail="Database unreachable.")


@router.get("/statement_generator", response_class=HTMLResponse, include_in_schema=False)
async def statement_generator_page(request: Request):
    return templates.TemplateResponse(request, "generate_statements.html")


@router.get("/terms", response_class=HTMLResponse, include_in_schema=False)
async def terms(request: Request):
    return templates.TemplateResponse(request, "legal_page.html", {
        "page_title": "Terms of Service", "active": "terms", "last_updated": "23 August 2026",
    })


@router.get("/privacy", response_class=HTMLResponse, include_in_schema=False)
async def privacy(request: Request):
    return templates.TemplateResponse(request, "legal_page.html", {
        "page_title": "Privacy Policy", "active": "privacy", "last_updated": "23 August 2026",
    })


@router.get("/download_examples", include_in_schema=False)
async def download_examples():
    """Public - lets a prospective user see exactly what a working
    template/spreadsheet pair looks like before they need a key."""
    examples_dir = os.path.join(STATIC_DIR, 'examples')
    zip_path = os.path.join(examples_dir, 'example_files.zip')

    if not os.path.exists(zip_path):
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(os.path.join(examples_dir, 'example_template.docx'), 'example_template.docx')
            zipf.write(os.path.join(examples_dir, 'example_data.xlsx'), 'example_data.xlsx')

    return FileResponse(zip_path, media_type='application/zip', filename='example_files.zip')


@router.post("/process_statement", response_model=ProcessStatementResponse, status_code=202,
             summary="Generate personalized PDF statements from a template + spreadsheet")
async def process_statement(
    template_file: UploadFile = File(..., description="Word template (.doc/.docx) with {{ field }} placeholders"),
    data_file: UploadFile = File(..., description="Spreadsheet (.xls/.xlsx) with one row per statement"),
    password_protection: bool = Form(False, description="Password-protect each PDF with that row's own ID"),
    client: dict = Depends(require_api_key),
):
    """Kicks off statement generation in the background and returns
    immediately - poll `/progress/{temp_id}` for status, then
    `/download_statement/{temp_id}` once complete."""
    if not template_file.filename:
        raise HTTPException(status_code=400, detail="Please choose a template file (.docx).")
    if not data_file.filename:
        raise HTTPException(status_code=400, detail="Please choose a data file (.xlsx).")
    if not allowed_template(template_file.filename):
        raise HTTPException(status_code=400, detail="Template file must be a .doc or .docx file.")
    if not allowed_data_file(data_file.filename):
        raise HTTPException(status_code=400, detail="Data file must be a .xls or .xlsx file.")

    if password_protection and not client['password_protection_allowed']:
        raise HTTPException(status_code=403, detail="Password protection isn't available on your current plan. Contact us to upgrade.")

    if client['monthly_job_limit'] is not None:
        try:
            used = await count_jobs_this_month(client['id'])
        except Exception as e:
            capture_service_error(e, where="process_statement.count_jobs_this_month", client_id=client['id'])
            raise HTTPException(status_code=503, detail="Service temporarily unavailable. Please try again shortly.")
        if used >= client['monthly_job_limit']:
            raise HTTPException(status_code=429, detail=(
                f"Monthly limit of {client['monthly_job_limit']} jobs reached for your plan. Contact us to upgrade."
            ))

    temp_id, expiry_time = generate_temp_id()
    output_folder = create_temp_folder(client['id'], temp_id)
    uploads_folder = os.path.join(output_folder, 'src')
    os.makedirs(uploads_folder, exist_ok=True)

    template_path = os.path.join(uploads_folder, secure_filename(template_file.filename))
    data_path = os.path.join(uploads_folder, secure_filename(data_file.filename))

    # UploadFile.read() is async (unlike Flask's synchronous FileStorage.save()) -
    # this is one of the real behavioral differences in the port, not just syntax.
    with open(template_path, 'wb') as f:
        f.write(await template_file.read())
    with open(data_path, 'wb') as f:
        f.write(await data_file.read())

    if client['max_rows_per_job'] is not None:
        try:
            row_count = count_data_rows(data_path)
        except Exception as e:
            shutil.rmtree(output_folder, ignore_errors=True)
            logger.warning(f"Could not read data file for row-count check: {e}")
            raise HTTPException(status_code=400, detail="Could not read the data file. Please check it's a valid .xls or .xlsx file.")

        if row_count > client['max_rows_per_job']:
            shutil.rmtree(output_folder, ignore_errors=True)
            raise HTTPException(status_code=400, detail=(
                f"Your plan allows up to {client['max_rows_per_job']} rows per job "
                f"(this file has {row_count}). Contact us to upgrade."
            ))
    else:
        row_count = None

    try:
        await create_job(temp_id, client['id'], row_count=row_count)
    except Exception as e:
        shutil.rmtree(output_folder, ignore_errors=True)
        capture_service_error(e, where="process_statement.create_job", client_id=client['id'], temp_id=temp_id)
        raise HTTPException(status_code=503, detail="Service temporarily unavailable. Please try again shortly.")

    TEMP_DOWNLOADS[temp_id] = {'folder': output_folder, 'expiry': expiry_time, 'owner': client['id']}
    set_progress(temp_id, "processing", "Queued...", 0)
    logger.info(f"Job {temp_id} started by client '{client['label']}' (id={client['id']})")

    asyncio.create_task(run_generation_job(
        temp_id, client['id'], client['label'], template_path, data_path, output_folder, password_protection
    ))

    return ProcessStatementResponse(
        temp_id=temp_id,
        message="Processing started.",
        status_url=f"/api/progress/{temp_id}",
        download_link=f"/api/download_statement/{temp_id}",
    )


@router.get("/progress/{temp_id}", response_model=ProgressResponse, summary="Check generation progress")
async def get_progress(temp_id: str, client: dict = Depends(require_api_key)):
    download_info = TEMP_DOWNLOADS.get(temp_id)
    # Same "unknown" response whether the id never existed or belongs to
    # someone else - this avoids confirming to a caller that a given temp_id
    # exists at all if it isn't theirs.
    if download_info is None or download_info.get('owner') != client['id']:
        raise HTTPException(status_code=404, detail={"state": "unknown", "message": "Unknown or expired ID.", "progress": 0})

    try:
        expected_folder = _safe_join(_client_root(client['id']), temp_id)
    except ValueError:
        raise HTTPException(status_code=404, detail={"state": "unknown", "message": "Unknown or expired ID.", "progress": 0})
    if os.path.realpath(download_info['folder']) != expected_folder:
        logger.error(f"Folder for {temp_id} is not under client {client['id']}'s root - refusing to report on it.")
        raise HTTPException(status_code=404, detail={"state": "unknown", "message": "Unknown or expired ID.", "progress": 0})

    task_status = PROGRESS_STATUS.get(temp_id)
    if task_status is None:
        raise HTTPException(status_code=404, detail={"state": "unknown", "message": "Unknown or expired ID.", "progress": 0})
    return ProgressResponse(**task_status)


@router.get("/download_statement/{temp_id}", summary="Download all generated PDFs as a zip")
async def download_all_files(temp_id: str, client: dict = Depends(require_api_key)):
    download_info = TEMP_DOWNLOADS.get(temp_id)
    if download_info is None or download_info.get('owner') != client['id']:
        raise HTTPException(status_code=404, detail="Invalid or expired download ID.")

    try:
        expected_folder = _safe_join(_client_root(client['id']), temp_id)
    except ValueError:
        raise HTTPException(status_code=404, detail="Invalid or expired download ID.")
    if os.path.realpath(download_info['folder']) != expected_folder:
        logger.error(f"Folder for {temp_id} is not under client {client['id']}'s root - refusing to serve it.")
        raise HTTPException(status_code=404, detail="Invalid or expired download ID.")

    task_status = PROGRESS_STATUS.get(temp_id)
    if task_status and task_status.get('state') != 'completed':
        raise HTTPException(status_code=409, detail="Files are not ready yet.")

    folder_path = download_info['folder']
    if not os.path.exists(folder_path):
        raise HTTPException(status_code=404, detail="Download folder not found.")

    try:
        pdf_files = [f for f in os.listdir(folder_path) if f.endswith('.pdf')]
        if not pdf_files:
            raise HTTPException(status_code=404, detail="No PDF files found to download.")

        zip_filename = f"statements_{temp_id}.zip"
        zip_filepath = os.path.join(folder_path, zip_filename)

        with zipfile.ZipFile(zip_filepath, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for pdf in pdf_files:
                zipf.write(os.path.join(folder_path, pdf), pdf)

        return FileResponse(zip_filepath, media_type='application/zip', filename=zip_filename)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating zip file for {temp_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Error creating download file: {str(e)}")